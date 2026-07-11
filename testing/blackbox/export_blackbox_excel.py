from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from data import COLUMNS, MODULE_SHEETS, cases_by_module, summary


OUTPUT = Path(__file__).resolve().parents[2] / "outputs" / "SWT_RMS_Blackbox_TestCases_98TC.xlsx"


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alt_fill = PatternFill("solid", fgColor="EAF3F8")
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = alt_fill

    widths = {
        "A": 14,
        "B": 14,
        "C": 36,
        "D": 34,
        "E": 36,
        "F": 42,
        "G": 42,
        "H": 12,
        "I": 18,
        "J": 12,
        "K": 24,
        "L": 12,
        "M": 14,
        "N": 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def add_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    info = summary()
    rows = [
        ("Document Code", "SWT_TC_v1.0"),
        ("Project", "Recruitment Management System (RMS)"),
        ("Created Date", info["created_date"]),
        ("Total Test Cases", info["total"]),
        ("Pass", info["pass"]),
        ("Fail", info["fail"]),
        ("Untested", info["untested"]),
        ("Modules", "EMP, CAN, ADM"),
    ]
    ws["A1"] = "BLACKBOX TEST CASE SPECIFICATION"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E78")
    for index, (label, value) in enumerate(rows, start=3):
        ws.cell(index, 1, label)
        ws.cell(index, 2, value)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 48


def add_module_sheet(wb, module):
    ws = wb.create_sheet(module)
    for col_idx, title in enumerate(COLUMNS, start=1):
        ws.cell(1, col_idx, title)
    for row_idx, case in enumerate(cases_by_module(module), start=2):
        for col_idx, title in enumerate(COLUMNS, start=1):
            ws.cell(row_idx, col_idx, case[title])
    style_sheet(ws)


def main():
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    add_cover(wb)
    for module in MODULE_SHEETS:
        add_module_sheet(wb, module)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
